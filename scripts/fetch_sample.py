"""Fetch Online Retail II and write the committed deterministic extract.

Governing decision: D-15 (docs/decisions/decision-register.md).
Source: Chen, D. (2012). Online Retail II [Dataset]. UCI Machine Learning
Repository. https://doi.org/10.24432/C5CG6D. License CC BY 4.0.

The window and row cap are constants by decision; this script takes no
arguments. Rerunning must produce a byte-identical CSV.
"""

from __future__ import annotations

import csv
import io
import shutil
import sys
import urllib.request
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

# D-15 constants: the window is code, not an argument.
SOURCE_URL = "https://archive.ics.uci.edu/static/public/502/online+retail+ii.zip"
SHEET_NAME = "Year 2009-2010"
WINDOW_START = datetime(2009, 12, 1)  # inclusive
WINDOW_END = datetime(2010, 1, 1)  # exclusive
# Fallback if December 2009 busts the size budget (edit both lines):
# WINDOW_START = datetime(2010, 2, 1); WINDOW_END = datetime(2010, 3, 1)
ROW_CAP = None  # no cap; shrink the window instead, invoices stay complete
MAX_BYTES = 5 * 1024 * 1024  # D-15 budget for the committed CSV

RAW_DIR = Path("data/raw")
RAW_ZIP = RAW_DIR / "online_retail_ii.zip"
XLSX_PATH = RAW_DIR / "online_retail_II.xlsx"
OUT_DIR = Path("data/samples/online_retail_ii")


def ensure_raw() -> Path:
    """Download and unzip the source workbook into gitignored data/raw/."""
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    if not RAW_ZIP.exists():
        print(f"downloading {SOURCE_URL}")
        req = urllib.request.Request(
            SOURCE_URL, headers={"User-Agent": "metricmine-fetch-sample/0.1"}
        )
        with urllib.request.urlopen(req) as resp, open(RAW_ZIP, "wb") as out:
            shutil.copyfileobj(resp, out)
    if not XLSX_PATH.exists():
        with zipfile.ZipFile(RAW_ZIP) as zf:
            print("zip contents:", zf.namelist())
            member = next(
                n for n in zf.namelist() if n.lower().endswith(".xlsx")
            )
            zf.extract(member, RAW_DIR)
            extracted = RAW_DIR / member
            if extracted != XLSX_PATH:
                extracted.rename(XLSX_PATH)
    return XLSX_PATH


def fmt(value: object) -> str:
    """Deterministic cell serialization, documented in the sample README."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def main() -> int:
    if ROW_CAP is not None:
        print("ERROR: ROW_CAP is not implemented; shrink WINDOW instead.")
        return 1
    xlsx_path = ensure_raw()
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    print("sheets:", wb.sheetnames)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: sheet {SHEET_NAME!r} not found; fix SHEET_NAME above.")
        return 1
    rows = wb[SHEET_NAME].iter_rows(values_only=True)

    raw_header = list(next(rows))
    while raw_header and raw_header[-1] is None:
        raw_header.pop()
    header = [str(h) for h in raw_header]
    print("header as landed:", header)
    date_idx = next(
        i
        for i, h in enumerate(header)
        if h.strip().lower().replace(" ", "") == "invoicedate"
    )

    months: Counter[str] = Counter()
    window_rows = []
    skipped = 0
    for row in rows:
        row = row[: len(header)]
        if all(v is None for v in row):
            continue
        ts = row[date_idx]
        if not isinstance(ts, datetime):
            skipped += 1
            continue
        months[ts.strftime("%Y-%m")] += 1
        if WINDOW_START <= ts < WINDOW_END:
            window_rows.append(row)
    wb.close()

    print(f"rows per month on sheet {SHEET_NAME!r}:")
    for month in sorted(months):
        print(f"  {month}: {months[month]}")
    if skipped:
        print(f"skipped {skipped} rows with non-datetime InvoiceDate")

    window_rows.sort(key=lambda r: tuple(fmt(v) for v in r))
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(header)
    for row in window_rows:
        writer.writerow([fmt(v) for v in row])
    data = buf.getvalue().encode("utf-8")

    mb = len(data) / (1024 * 1024)
    print(
        f"window {WINDOW_START:%Y-%m-%d}..{WINDOW_END:%Y-%m-%d}: "
        f"{len(window_rows)} rows, {mb:.2f} MB"
    )
    if len(data) > MAX_BYTES:
        print("ERROR: over the D-15 5 MB budget; choose a smaller WINDOW.")
        return 1
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_csv = OUT_DIR / f"online_retail_ii_{WINDOW_START:%Y-%m}.csv"
    out_csv.write_bytes(data)
    print(f"OK wrote {out_csv}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
